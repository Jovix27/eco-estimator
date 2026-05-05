"""
Export API — generate Excel workbooks for BOQ and estimates.
Requires openpyxl: pip install openpyxl
"""

from __future__ import annotations
import io
from typing import List

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from api.projects import _get_project, _get_boq, _get_estimate

router = APIRouter(prefix="/export", tags=["Export"])


def _fmt(val: float) -> str:
    return f"₹{val:,.0f}"


@router.get("/projects/{project_id}/excel")
async def export_excel(project_id: str):
    """Export project BOQ + Estimate to a formatted Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed — run: pip install openpyxl")

    project = _get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    boq_items = _get_boq(project_id)
    estimate = _get_estimate(project_id)

    wb = openpyxl.Workbook()

    # ── Styles ───────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1A3C34")
    trade_fill  = PatternFill("solid", fgColor="2D6A4F")
    total_fill  = PatternFill("solid", fgColor="081C15")
    alt_fill    = PatternFill("solid", fgColor="F0FDF4")

    bold_white  = Font(bold=True, color="FFFFFF", size=10)
    bold_green  = Font(bold=True, color="10B981", size=10)
    bold_dark   = Font(bold=True, color="1A3C34", size=10)
    normal      = Font(size=9)

    thin = Side(border_style="thin", color="D1FAE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")

    # ─────────────────── Sheet 1: PROJECT INFO ───────────────────────────
    ws_info = wb.active
    ws_info.title = "Project Info"
    ws_info.sheet_view.showGridLines = False

    info_rows = [
        ("ECO ESTIMATOR — PROJECT SUMMARY", ""),
        ("", ""),
        ("Project Name",    project.get("name", "")),
        ("Location",        project.get("location", "")),
        ("Building Type",   project.get("building_type", "").title()),
        ("Quality Grade",   project.get("quality", "").title()),
        ("Total Area",      f"{project.get('total_area_sqft', 0):,.0f} sqft"),
        ("No. of Floors",   project.get("num_floors", 1)),
        ("Created",         str(project.get("created_at", ""))[:10]),
    ]

    for r_idx, (label, value) in enumerate(info_rows, 1):
        cell_a = ws_info.cell(row=r_idx, column=1, value=label)
        cell_b = ws_info.cell(row=r_idx, column=2, value=value)
        if r_idx == 1:
            cell_a.font = Font(bold=True, color="10B981", size=14)
            ws_info.merge_cells(f"A1:D1")
        else:
            cell_a.font = bold_dark
            cell_b.font = normal

    ws_info.column_dimensions["A"].width = 22
    ws_info.column_dimensions["B"].width = 35

    # ─────────────────── Sheet 2: BOQ ────────────────────────────────────
    ws_boq = wb.create_sheet("BOQ — Item Rate")
    ws_boq.sheet_view.showGridLines = False

    boq_headers = ["Item No.", "Trade", "Description", "Unit", "Qty", "Rate (₹)", "Amount (₹)", "Material (₹)", "Labour (₹)", "Equip. (₹)"]
    col_widths   = [10, 18, 55, 8, 10, 12, 14, 13, 13, 12]

    for col, (h, w) in enumerate(zip(boq_headers, col_widths), 1):
        cell = ws_boq.cell(row=1, column=col, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws_boq.column_dimensions[get_column_letter(col)].width = w

    ws_boq.row_dimensions[1].height = 30

    row = 2
    current_trade = None
    trade_start = 2

    grouped: dict[str, List[dict]] = {}
    for item in boq_items:
        t = item["trade"]
        grouped.setdefault(t, []).append(item)

    for trade, items in grouped.items():
        # Trade header row
        trade_cell = ws_boq.cell(row=row, column=1, value=trade.upper())
        ws_boq.merge_cells(f"A{row}:J{row}")
        trade_cell.font = bold_white
        trade_cell.fill = trade_fill
        trade_cell.alignment = left
        trade_cell.border = border
        ws_boq.row_dimensions[row].height = 20
        row += 1

        trade_total = 0.0
        for i_idx, item in enumerate(items):
            fill = alt_fill if i_idx % 2 == 0 else None
            values = [
                item["item_no"],
                item["trade"],
                item["description"],
                item["unit"],
                round(item["quantity"], 3),
                round(item["rate"], 2),
                round(item["amount"], 2),
                round(item["material_rate"] * item["quantity"], 2),
                round(item["labor_rate"] * item["quantity"], 2),
                round(item["equipment_rate"] * item["quantity"], 2),
            ]
            for col, val in enumerate(values, 1):
                c = ws_boq.cell(row=row, column=col, value=val)
                c.font = normal
                c.alignment = center if col in (1, 4, 5) else (left if col == 3 else right)
                c.border = border
                if fill:
                    c.fill = fill
            ws_boq.row_dimensions[row].height = 40
            trade_total += item["amount"]
            row += 1

        # Trade subtotal
        sub_cell = ws_boq.cell(row=row, column=3, value=f"{trade} Sub-Total")
        ws_boq.cell(row=row, column=7, value=round(trade_total, 2))
        for col in range(1, 11):
            c = ws_boq.cell(row=row, column=col)
            c.font = bold_green
            c.fill = PatternFill("solid", fgColor="ECFDF5")
            c.border = border
        row += 1

    # Grand total row
    total = round(sum(i["amount"] for i in boq_items), 2)
    ws_boq.cell(row=row, column=3, value="TOTAL DIRECT COST").font = bold_white
    ws_boq.cell(row=row, column=7, value=total).font = bold_white
    for col in range(1, 11):
        c = ws_boq.cell(row=row, column=col)
        c.fill = total_fill
        c.border = border
        c.alignment = right

    # ─────────────────── Sheet 3: ESTIMATE ───────────────────────────────
    if estimate:
        ws_est = wb.create_sheet("Cost Estimate")
        ws_est.sheet_view.showGridLines = False

        est_rows = [
            ("COST ESTIMATE SUMMARY", ""),
            ("", ""),
            ("Estimation Method",    estimate.get("method", "Item Rate Method")),
            ("", ""),
            ("1. Direct Cost (BOQ)",         _fmt(estimate.get("direct_cost", 0))),
            ("2. Site Overhead",             f"{estimate.get('overhead_pct',0)}% = {_fmt(estimate.get('overhead_amount',0))}"),
            ("3. Contractor Profit",         f"{estimate.get('profit_pct',0)}% = {_fmt(estimate.get('profit_amount',0))}"),
            ("4. Contingency",               f"{estimate.get('contingency_pct',0)}% = {_fmt(estimate.get('contingency_amount',0))}"),
            ("",                             ""),
            ("Sub-Total",                    _fmt(estimate.get("sub_total", 0))),
            ("5. GST",                       f"{estimate.get('gst_pct',0)}% = {_fmt(estimate.get('gst_amount',0))}"),
            ("",                             ""),
            ("GRAND TOTAL",                  _fmt(estimate.get("grand_total", 0))),
            ("",                             ""),
            ("Cost per Sqft (Direct)",       f"₹{estimate.get('cost_per_sqft',0):,.0f}"),
            ("Cost per Sqft (Grand Total)",  f"₹{estimate.get('grand_total_per_sqft',0):,.0f}"),
            ("",                             ""),
            ("Material Cost",                _fmt(estimate.get("total_material_cost", 0))),
            ("Labour Cost",                  _fmt(estimate.get("total_labor_cost", 0))),
            ("Equipment Cost",               _fmt(estimate.get("total_equipment_cost", 0))),
        ]

        for r_idx, (label, value) in enumerate(est_rows, 1):
            cell_a = ws_est.cell(row=r_idx, column=1, value=label)
            cell_b = ws_est.cell(row=r_idx, column=2, value=value)

            if r_idx == 1:
                cell_a.font = Font(bold=True, color="10B981", size=14)
            elif label in ("GRAND TOTAL",):
                cell_a.font = bold_white; cell_a.fill = total_fill
                cell_b.font = bold_white; cell_b.fill = total_fill
            elif label in ("Sub-Total",):
                cell_a.font = bold_green
                cell_b.font = bold_green
            else:
                cell_a.font = bold_dark
                cell_b.font = normal

        ws_est.column_dimensions["A"].width = 35
        ws_est.column_dimensions["B"].width = 28

    # ── Stream the workbook ──────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"EcoEstimator_{project.get('name','Project').replace(' ','_')}_BOQ.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
